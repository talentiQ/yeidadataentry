from __future__ import annotations

import json
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from shutil import copyfile
from typing import Dict, List

from openpyxl import load_workbook

# =========================================================
# TEMPLATE FILE
# =========================================================

import os

BASE_DIR      = Path(__file__).resolve().parent
TEMPLATE_FILE = BASE_DIR / "data" / "RPS_10_MASTER.xlsx"

TEMPLATE_FORMULA_MAX_ROW = 10050

_VALIDATION_FORMULAS: Dict[str, str] = {
    "AU": (
        '=IF(N{R}="","",IF(AND('
        'LEN(TRIM(N{R}))=10,'
        'CODE(UPPER(MID(N{R},1,1)))>=65,CODE(UPPER(MID(N{R},1,1)))<=90,'
        'CODE(UPPER(MID(N{R},2,1)))>=65,CODE(UPPER(MID(N{R},2,1)))<=90,'
        'CODE(UPPER(MID(N{R},3,1)))>=65,CODE(UPPER(MID(N{R},3,1)))<=90,'
        'UPPER(MID(N{R},4,1))="P",'
        'CODE(UPPER(MID(N{R},5,1)))>=65,CODE(UPPER(MID(N{R},5,1)))<=90,'
        'ISNUMBER(VALUE(MID(N{R},6,1))),ISNUMBER(VALUE(MID(N{R},7,1))),'
        'ISNUMBER(VALUE(MID(N{R},8,1))),ISNUMBER(VALUE(MID(N{R},9,1))),'
        'CODE(UPPER(MID(N{R},10,1)))>=65,CODE(UPPER(MID(N{R},10,1)))<=90'
        '),"✔ Valid PAN","✘ Invalid PAN"))'
    ),
    "AV": '=IF(N{R}="","",IF(COUNTIF($N$4:$N$10000,N{R})>1,"✗ DUPLICATE PAN","✓ Unique"))',
    "AW": '=IF(AA{R}="","",IF(COUNTIF($AA$4:$AA$10000,AA{R})>1,"✗ DUPLICATE Mobile","✓ Unique"))',
    "AX": '=IF(AP{R}="","",IF(ISNUMBER(MATCH(VALUE(AP{R}),{162,183,184,200,223,290},0)),"✔ Valid Size","✘ Invalid Size"))',
    "AY": '=IF(M{R}="","",IF(OR(UPPER(TRIM(M{R}))="GEN",UPPER(TRIM(M{R}))="SC/ST",UPPER(TRIM(M{R}))="SC",UPPER(TRIM(M{R}))="ST"),"✔ Valid Cat","✘ Invalid Cat"))',
    "AZ": '=IF(AC{R}="","",IF(UPPER(TRIM(AC{R}))="AUTO DEBIT","✔ Valid IFSC",IF(LEN(TRIM(AF{R}))=11,"✔ Valid IFSC","✘ Invalid IFSC")))',
    "BA": '=IF(AA{R}="","",IF(AND(ISNUMBER(AA{R}),LEN(TEXT(AA{R},"0"))=10),"✔ Valid Mob","✘ Invalid Mob"))',
    "BB": '=IF(D{R}="","",IF(AND(LEN(TRIM(D{R}))=12,EXACT(UPPER(TRIM(D{R})),TRIM(D{R}))),"✔ Valid RPS","✘ RPS Must Be 12 Chars"))',
    "BC": (
        '=IF(AND(N{R}="",AP{R}=""),"",IF('
        'SUMPRODUCT(--(ISNUMBER(SEARCH({"DUPLICATE","ERROR","✗","Invalid"},AU{R}:BB{R}))))>0,'
        '"✗ ERRORS – CHECK","✓ READY TO SUBMIT"))'
    ),
}

# =========================================================
# HELPERS
# =========================================================

def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _find_next_empty_row(ws) -> int:
    """Return the first row >= 4 where both col D and col E are blank."""
    row = 4
    while True:
        if ws[f"D{row}"].value in (None, "") and ws[f"E{row}"].value in (None, ""):
            return row
        row += 1


def _inject_validation_formulas(ws, row: int) -> None:
    """Write AU–BC validation formulas for rows beyond the template pre-fill."""
    if row <= TEMPLATE_FORMULA_MAX_ROW:
        return
    for col, tmpl in _VALIDATION_FORMULAS.items():
        ws[f"{col}{row}"] = tmpl.replace("{R}", str(row))


# =========================================================
# APPEND TO EXCEL
# =========================================================

def append_to_excel(row: Dict, output_path: Path) -> int:
    """
    Append one candidate row to the master Excel.
    Creates the file from template on first call.
    Returns the row number written.
    """
    ensure_parent(output_path)
    if not output_path.exists():
        copyfile(TEMPLATE_FILE, output_path)

    wb = load_workbook(output_path)
    ws = wb.active
    assert ws is not None, "Workbook has no active sheet"

    next_row = _find_next_empty_row(ws)

    ws[f"A{next_row}"] = next_row - 3
    ws[f"B{next_row}"] = row.get("lot_no", "")
    ws[f"C{next_row}"] = row.get("file_no", "")
    ws[f"D{next_row}"] = row.get("rps_no", "")
    ws[f"E{next_row}"] = row.get("receipt_no", "")
    ws[f"F{next_row}"] = row.get("sourcing_branch", "")
    ws[f"G{next_row}"] = row.get("first_name", "")
    ws[f"H{next_row}"] = row.get("middle_name", "")
    ws[f"I{next_row}"] = row.get("last_name", "")
    ws[f"J{next_row}"] = row.get("dob", "")
    ws[f"K{next_row}"] = row.get("gender", "")
    ws[f"L{next_row}"] = row.get("marital_status", "")
    ws[f"M{next_row}"] = row.get("category", "")
    ws[f"N{next_row}"] = row.get("pan_no", "")
    ws[f"O{next_row}"] = row.get("religion", "")
    ws[f"P{next_row}"] = row.get("current_residence", "")
    ws[f"Q{next_row}"] = row.get("qualification", "")
    ws[f"R{next_row}"] = row.get("profession", "")
    ws[f"S{next_row}"] = row.get("address_1", "")
    ws[f"T{next_row}"] = row.get("address_2", "")
    ws[f"U{next_row}"] = row.get("address_3", "")
    ws[f"V{next_row}"] = row.get("city", "")
    ws[f"W{next_row}"] = row.get("state", "")
    ws[f"X{next_row}"] = row.get("zip_code", "")
    ws[f"Y{next_row}"] = row.get("phone_1", "")
    ws[f"Z{next_row}"] = row.get("phone_2", "")
    ws[f"AA{next_row}"] = row.get("mobile", "")
    ws[f"AB{next_row}"] = row.get("email", "")
    ws[f"AC{next_row}"] = row.get("repayment_mode", "")
    ws[f"AD{next_row}"] = row.get("account_no", "")
    ws[f"AE{next_row}"] = row.get("micr", "")
    ws[f"AF{next_row}"] = row.get("ifsc", "")
    ws[f"AG{next_row}"] = row.get("bank_name", "")
    ws[f"AH{next_row}"] = row.get("account_type", "")
    ws[f"AI{next_row}"] = row.get("id_proof_no", "")
    ws[f"AJ{next_row}"] = row.get("id_expiry", "")
    ws[f"AK{next_row}"] = row.get("id_doc_type", "")
    ws[f"AL{next_row}"] = row.get("address_proof_no", "")
    ws[f"AM{next_row}"] = row.get("address_expiry", "")
    ws[f"AN{next_row}"] = row.get("address_doc_type", "")
    ws[f"AO{next_row}"] = row.get("sc_st_flag", "")
    ws[f"AP{next_row}"] = row.get("plot_size", "")
    ws[f"AQ{next_row}"] = row.get("asset_cost", "")
    ws[f"AR{next_row}"] = row.get("amt_financed", "")
    ws[f"AS{next_row}"] = 5900  # PF Amount fixed
    try:
        interest = round(float(row.get("amt_financed") or 0) * 0.10)
    except Exception:
        interest = 0
    ws[f"AT{next_row}"] = interest

    _inject_validation_formulas(ws, next_row)
    wb.save(output_path)
    return next_row


# =========================================================
# READ ALL ROWS
# =========================================================

def read_all_rows_from_excel(path: Path) -> List[Dict]:
    """
    Read all data rows from the master Excel as a list of dicts.
    Returns an empty list if the file does not exist.
    """
    if not path.exists():
        return []

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    assert ws is not None, "Workbook has no active sheet"

    DATA_COLS = {
        "A": "serial_no",       "B": "lot_no",          "C": "file_no",
        "D": "rps_no",          "E": "receipt_no",       "F": "sourcing_branch",
        "G": "first_name",      "H": "middle_name",      "I": "last_name",
        "J": "dob",             "K": "gender",           "L": "marital_status",
        "M": "category",        "N": "pan_no",           "O": "religion",
        "P": "current_residence","Q": "qualification",   "R": "profession",
        "S": "address_1",       "T": "address_2",        "U": "address_3",
        "V": "city",            "W": "state",            "X": "zip_code",
        "Y": "phone_1",         "Z": "phone_2",          "AA": "mobile",
        "AB": "email",          "AC": "repayment_mode",  "AD": "account_no",
        "AE": "micr",           "AF": "ifsc",            "AG": "bank_name",
        "AH": "account_type",   "AI": "id_proof_no",     "AJ": "id_expiry",
        "AK": "id_doc_type",    "AL": "address_proof_no","AM": "address_expiry",
        "AN": "address_doc_type","AO": "sc_st_flag",     "AP": "plot_size",
        "AQ": "asset_cost",     "AR": "amt_financed",    "AS": "pf_amount",
        "AT": "interest_amt",
    }

    rows_out = []
    for r in range(4, ws.max_row + 1):
        data = {key: ws[f"{col}{r}"].value for col, key in DATA_COLS.items()}
        # Skip completely blank rows
        if data["rps_no"] in (None, "") and data["receipt_no"] in (None, ""):
            continue
        rows_out.append(data)
    return rows_out


# =========================================================
# XML EXPORT
# =========================================================

def write_xml_from_excel(xlsx_path: Path, xml_path: Path) -> None:
    """Rebuild the XML export from the current Excel data."""
    rows = read_all_rows_from_excel(xlsx_path)
    root = ET.Element("RPSApplications")
    root.set("generated_at", datetime.now().isoformat())
    root.set("count", str(len(rows)))

    for idx, row_data in enumerate(rows, start=1):
        app_el = ET.SubElement(root, "Application")
        app_el.set("row", str(idx))
        for key, value in row_data.items():
            child = ET.SubElement(app_el, key)
            child.text = "" if value is None else str(value)

    ensure_parent(xml_path)
    tree = ET.ElementTree(root)
    ET.indent(tree, space="  ")
    tree.write(xml_path, encoding="utf-8", xml_declaration=True)


# =========================================================
# JSON LOG
# =========================================================

def append_json_log(row: Dict, json_path: Path) -> None:
    """Append one row to the running JSON log file."""
    ensure_parent(json_path)
    existing: List[Dict] = []
    if json_path.exists():
        try:
            existing = json.loads(json_path.read_text(encoding="utf-8"))
        except Exception:
            existing = []
    existing.append({"saved_at": datetime.now().isoformat(), **row})
    json_path.write_text(
        json.dumps(existing, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


# =========================================================
# PYTHON VALIDATION (fallback — mirrors the Excel AU–BC formulas)
# =========================================================

def _compute_validation(data: Dict) -> Dict:
    pan      = str(data.get("pan_no") or "").strip().upper()
    mobile   = data.get("mobile")
    plot     = data.get("plot_size")
    category = str(data.get("category") or "").strip().upper()
    ifsc     = str(data.get("ifsc") or "").strip()
    repay    = str(data.get("repayment_mode") or "").strip().upper()
    rps      = str(data.get("rps_no") or "").strip()

    pan_ok = bool(re.match(r'^[A-Z]{3}P[A-Z]\d{4}[A-Z]$', pan)) if pan else None
    au = "✔ Valid PAN" if pan_ok else ("✘ Invalid PAN" if pan else "")
    av = "⚠ Dup check N/A"
    aw = "⚠ Dup check N/A"

    valid_sizes = {162, 183, 184, 200, 223, 290}
    try:
        ax = "✔ Valid Size" if int(float(str(plot))) in valid_sizes else "✘ Invalid Size"
    except Exception:
        ax = "" if not plot else "✘ Invalid Size"

    ay = "✔ Valid Cat" if category in {"GEN", "SC/ST", "SC", "ST"} else ("✘ Invalid Cat" if category else "")

    if not repay and not ifsc:
        az = ""
    elif repay == "AUTO DEBIT":
        az = "✔ Valid IFSC"
    else:
        az = "✔ Valid IFSC" if len(ifsc) == 11 else "✘ Invalid IFSC"

    try:
        mob_str = str(int(float(str(mobile)))) if mobile is not None else ""
        ba = "✔ Valid Mob" if len(mob_str) == 10 else "✘ Invalid Mob"
    except Exception:
        ba = "" if not mobile else "✘ Invalid Mob"

    bb = (
        "✔ Valid RPS" if (len(rps) == 12 and rps == rps.upper())
        else ("✘ RPS Must Be 12 Chars" if rps else "")
    )

    errors = [
        v for v in [au, av, aw, ax, ay, az, ba, bb]
        if any(kw in str(v) for kw in ("DUPLICATE", "ERROR", "✗", "Invalid", "✘"))
    ]
    bc = (
        "✓ READY TO SUBMIT" if (pan or plot) and not errors
        else ("✗ ERRORS – CHECK" if errors else "")
    )

    return {
        "pan_format":       au,
        "pan_duplicate":    av,
        "mobile_duplicate": aw,
        "plot_size_check":  ax,
        "category_check":   ay,
        "ifsc_check":       az,
        "mobile_format":    ba,
        "rps_no_check":     bb,
        "overall_status":   bc,
    }
